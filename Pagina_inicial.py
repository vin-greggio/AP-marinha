import streamlit as st
import hashlib
import json
import os
import pandas as pd
import polars as pl
import matplotlib.pyplot as plt
import matplotlib
import plotly_express as px
import openpyxl as ox
from streamlit_option_menu import option_menu
from supabase import create_client, Client
from datetime import datetime

#INICIO DATAFRAMES

matplotlib.use('TkAgg')

@st.cache_resource
def init_connection():
    url = "https://fudwcyqizyszfqgwvmik.supabase.co"
    key = 'eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6ImZ1ZHdjeXFpenlzemZxZ3d2bWlrIiwicm9sZSI6ImFub24iLCJpYXQiOjE3MjU5ODU2NDIsImV4cCI6MjA0MTU2MTY0Mn0.OfjRY2WeMZLlxgMJzYuTOLM390FWBrnOY9kGKp3oYgA'
    return create_client(url, key)

supabase = init_connection()

#INICIO DAS FUNÇOES
#########################################################################################
def pre_fill_excel(template_file, supabase_client):
    # Obter dados do Supabase
    response = supabase_client.table('NOME_DA_TABELA').select('*').execute()
    data = response.data
    
    # Carregar o template do Excel
    wb = ox.load_workbook(template_file)
    ws = wb.active
    
    # Preencher planilha com dados (exemplo: preencher com o primeiro registro)
    for row in data:
        ws['A1'] = row['coluna1']
        ws['B2'] = row['coluna2']
        break  # Exemplo simples, ajuste conforme necessário

    # Salvar o arquivo pré-preenchido em memória

def hash_password(password):
    return hashlib.sha256(password.encode()).hexdigest()

def load_users():
    if os.path.exists("users.json"):
        with open("users.json", "r") as file:
            return json.load(file)
    return {}

def save_users(users):
    with open("users.json", "w") as file:
        json.dump(users, file)

users_db = load_users()

def authenticate(username, password):
    if username in users_db and users_db[username]['password'] == hash_password(password):
        return True
    return False

def create_user(username, password, role):
    if username in users_db:
        return False
    users_db[username] = {"password": hash_password(password), "role": role}
    save_users(users_db)
    return True

def change_password(username, old_password, new_password):
    if authenticate(username, old_password):
        users_db[username]['password'] = hash_password(new_password)
        save_users(users_db)
        return True
    return False

def has_permission(username, role):
    return users_db.get(username, {}).get("role") == role

#FIM DAS FUNCOES
####################################################################################

st.title("Sistema de Planilhas AP Marinha")
st.sidebar.image('imgmar.jpeg')
st.sidebar.header("Login")
with st.sidebar:
    selected = option_menu(
        "Menu",
        ["Login", "Empréstimo",'Dashboard','Upload'],
        icons=["key", "calculator",'chart','']
    )
if 'authenticated' not in st.session_state:
    st.session_state['authenticated'] = False
login_button = 0
if not st.session_state['authenticated']:
    with st.form('Login'):
        st.write('Insira suas credenciais')
        username = st.text_input('Usuário')
        password = st.text_input("Senha", type="password")
        login_button = st.form_submit_button("Login")


if login_button:
    if authenticate(username, password):
        st.sidebar.success(f"Bem-vindo {username}!")
        st.session_state['authenticated'] = True
        auth = 1
        st.session_state['username'] = username
        st.session_state['role'] = users_db[username]['role']
    else:
        st.sidebar.error("Usuário ou senha incorretos.")
####################################################################################


# COMEÇA AQUI A PARTIR DE LOGADO


####################################################################################
if 'authenticated' in st.session_state and st.session_state['authenticated']:
    
    st.write(f"Usuário logado: {st.session_state['username']} ({st.session_state['role']})")


    #CASO EDITOR:
    if st.session_state['role'] == 'editor':

        if selected=='Dashboard':
            saude = pd.DataFrame(supabase.table('receitasxdespesas').select("*").execute().data)
            desocupados = pd.DataFrame(supabase.table('desocupados').select("*").execute().data)
            st.subheader('Dashboard')
            saude_tab,desocupados_tab = st.tabs(['Saúde Financeira','Desocupados'])
            with saude_tab:
                st.write('Saldo médio: ',saude['SALDO'].mean())
                pos_neg = st.selectbox(label='Selecione o gráfico desejado',options=['Condomínios com saldo positivo','Condomínios com saldo negativo'])
                if(pos_neg=='Condomínios com saldo positivo'):
                    fig_saude = px.bar(saude[saude['SALDO']>0].sort_values('SALDO'),x='CONDOMINIO', y='SALDO',width=1000)
                    st.plotly_chart(fig_saude)
                elif(pos_neg=='Condomínios com saldo negativo'):
                    fig_saude = px.bar(saude[saude['SALDO']<=0].sort_values('SALDO'),x='CONDOMINIO', y='SALDO',width=1000)
                    st.plotly_chart(fig_saude)
            with desocupados_tab:
                desocupados['COND'] = desocupados['COND'].str.replace('.', '', regex=False).str.replace(',', '.', regex=False)
                desocupados['COND'] = desocupados['COND'].str.slice(3).astype(float)
                st.write('Total de apartamentos desocupados: ',len(desocupados))
                st.write('Soma de valores condominiais: ',sum(desocupados['COND']), ' reais')
                st.write('Média de valores condominiais: ', desocupados['COND'].mean().round(2), ' reais')
                condo_selecionado = st.selectbox(label='Selecione o gráfico desejado',options=['Águas Claras',
                'VNAVI','SHCES','SHCGN','SQS','SHIGS','Guara','Outros'])

                print(desocupados.columns)
                desocupados = desocupados[desocupados['SITUACAO']=='DESOCUPADO']
                if condo_selecionado=='VNAVI':
                    fig_des = px.bar(desocupados[desocupados['PNR'].str.startswith('VNAVI')].sort_values('COND'),y='PNR',x='COND',orientation='h',height=1000)
                    fig_des.update_layout(bargap=0.1)
                if condo_selecionado=='Águas Claras':
                    fig_des = px.bar(desocupados[desocupados['PNR'].str.startswith('Águas Claras')].sort_values('COND'),y='PNR',x='COND',orientation='h')
                    fig_des.update_layout(bargap=0.1)
                if condo_selecionado=='SHCES':
                    fig_des = px.bar(desocupados[desocupados['PNR'].str.startswith('SHCES')].sort_values('COND'),y='PNR',x='COND',orientation='h')
                    fig_des.update_layout(bargap=0.1)
                if condo_selecionado=='SHCGN':
                    fig_des = px.bar(desocupados[desocupados['PNR'].str.startswith('SHCGN')].sort_values('COND'),y='PNR',x='COND',orientation='h')
                    fig_des.update_layout(bargap=0.1)
                if condo_selecionado=='SHIGS':
                    fig_des = px.bar(desocupados[desocupados['PNR'].str.startswith('SHIGS')].sort_values('COND'),y='PNR',x='COND',orientation='h')
                    fig_des.update_layout(bargap=0.1)
                if condo_selecionado=='SQS':
                    fig_des = px.bar(desocupados[desocupados['PNR'].str.startswith('SQS')].sort_values('COND'),y='PNR',x='COND',orientation='h')
                    fig_des.update_layout(bargap=0.1)
                if condo_selecionado=='Guara':
                    fig_des = px.bar(desocupados[desocupados['PNR'].str.startswith('Guara')].sort_values('COND'),y='PNR',x='COND',orientation='h')
                    fig_des.update_layout(bargap=0.1)
                if condo_selecionado=='Outros':
                    fig_des = px.bar(desocupados[~desocupados['PNR'].str.startswith(('Águas Claras',
                'VNAVI','SHCES','SHCGN','SQS','SHIGS','Guara'))].sort_values('COND'),y='PNR',x='COND',orientation='h')
                    fig_des.update_layout(bargap=0.1)

                st.plotly_chart(fig_des)
        if selected=='Upload':
            restituicoes_file = st.file_uploader("Insira o CSV da planilha restituições", type="csv")
            desocupados_file = st.file_uploader("Insira o CSV da planilha desocupados", type="csv")
            isolados_file = st.file_uploader("Insira o CSV da planilha isolados", type="csv")
            taxaextra_file = st.file_uploader("Insira o CSV da planilha taxa extra", type="csv")

            if restituicoes_file is not None:
                restituicoes_df = pd.read_csv(restituicoes_file)
                st.write("Insira o arquivo CSV:", restituicoes_df)
                
                table_name = "restituicoes"
                
                data_to_insert = restituicoes_df.to_dict(orient='records')

                response = supabase.table(table_name).insert(data_to_insert).execute()
                st.write(response)
            
            if desocupados_file is not None:
                desocupados_df = pd.read_csv(desocupados_file)
                st.write("Insira o arquivo CSV:", desocupados_df)
                
                table_name = "desocupados"
                
                data_to_insert = desocupados_df.to_dict(orient='records')

                response = supabase.table(table_name).insert(data_to_insert).execute()
                st.write(response)
            
            if isolados_file is not None:
                isolados_df = pd.read_csv(isolados_file)
                st.write("Insira o arquivo CSV:", isolados_df)
                
                table_name = "isolados"
                
                data_to_insert = isolados_df.to_dict(orient='records')

                response = supabase.table(table_name).insert(data_to_insert).execute()
                st.write(response)

            if taxaextra_file is not None:
                taxaextra_df = pd.read_csv(taxaextra_file)
                st.write("Insira o arquivo CSV:", taxaextra_df)
                
                table_name = "taxa_extra"
                
                data_to_insert = taxaextra_df.to_dict(orient='records')

                response = supabase.table(table_name).insert(data_to_insert).execute()
                st.write(response)
        
        elif selected=='Alterar Senha':
                st.subheader("Alterar senha")
                old_password = st.text_input("Senha antiga", type="password")
                new_password = st.text_input("Nova senha", type="password")
                change_password_button = st.button("Alterar senha")

                if change_password_button:
                    if change_password(st.session_state['username'], old_password, new_password):
                        st.success("Senha alterada com sucesso!")
                    else:
                        st.error("Senha antiga incorreta.")
        elif selected == 'Criar Usuário':
            st.subheader("Criar novo usuário")
            new_username = st.text_input("Novo usuário")
            new_password = st.text_input("Primeira Senha", type="password")
            new_role = st.selectbox("Função", ["editor", "viewer"])
            create_user_button = st.button("Criar usuário")

            if create_user_button:
                if create_user(new_username, new_password, new_role):
                    st.success("Usuário criado com sucesso!")
                else:
                    st.error("Usuário já existe.")
    #CASO NÃO EDITOR:
    else:
        with st.sidebar:
            selected = option_menu(
                    "Menu",
                    ["Planilhas", 'Empréstimo', 'Alterar Senha', 'Criar Usuário'],
                    icons=['info', 'calculator'],)
            
        if selected=='Alterar Senha':
            st.subheader("Alterar senha")
            old_password = st.text_input("Senha antiga", type="password")
            new_password = st.text_input("Nova senha", type="password")
            change_password_button = st.button("Alterar senha")

            if change_password_button:
                if change_password(st.session_state['username'], old_password, new_password):
                    st.success("Senha alterada com sucesso!")
                else:
                    st.error("Senha antiga incorreta.")
else:
    st.write("Por favor, faça login para continuar.")

if selected == "Empréstimo":

    # Função para buscar empréstimos existentes no banco de dados
    def load_emprestimos():
        try:
            response = supabase.table("emprestimos").select("*").execute()
            return response.data if response.data else []
        except Exception as e:
            st.error(f"Erro ao carregar empréstimos: {e}")
            return []
        
    # Função para buscar parcelas existentes no banco de dados
    def load_parcelas(emprestimo_id):
        try:
            response = supabase.table("parcelas").select("*").eq("emprestimo_id", emprestimo_id).execute()
            return response.data if response.data else []
        except Exception as e:
            st.error(f"Erro ao carregar parcelas: {e}")
            return []
    
    # Função para salvar um novo empréstimo
    def save_emprestimo(emprestimo_data):
        try:
            response = supabase.table("emprestimos").insert([emprestimo_data]).execute()
            if response.data:
                return response.data  # Retorna os dados do novo empréstimo
            else:
                st.error("Erro ao salvar o empréstimo. Verifique os dados.")
                return None
        except Exception as e:
            st.error(f"Erro ao salvar empréstimo: {e}")
            return None

    # Função para salvar parcelas
    def save_parcelas(parcelas_data):
        try:
            # Verifica se cada parcela tem o campo 'valor_parcela'
            for parcela in parcelas_data:
                if 'valor_parcela' not in parcela:
                    parcela['valor_parcela'] = 0.0  # Atribui um valor padrão se não existir

            response = supabase.table("parcelas").insert(parcelas_data).execute()
            return response.data is not None
        except Exception as e:
            st.error(f"Erro ao salvar parcelas: {e}")
            return False

    # Função para deletar parcela
    def delete_parcela(parcela_id):
        try:
            response = supabase.table("parcelas").delete().eq('id', parcela_id).execute()
            return response.data is not None
        except Exception as e:
            st.error(f"Erro ao deletar parcela: {e}")
            return False

    # Função para atualizar parcela
    def update_parcela(parcela_id, updated_data):
        try:
            response = supabase.table('parcelas').update(updated_data).eq('id', parcela_id).execute()
            return response.data is not None
        except Exception as e:
            st.error(f"Erro ao atualizar a parcela: {e}")
            return False

    def marcar_parcela_como_pago(parcela_id):
        try:
            response = supabase.table('parcelas').update({"status": "Pago"}).eq("id", parcela_id).execute()
            return response.data is not None
        except Exception as e:
            st.error(f"Erro ao marcar a parcela como paga: {e}")
            return False

    # Atualiza o valor total do empréstimo
    def update_valor_total_emprestimo(emprestimo_id, novo_valor_total):
        try:
            response = supabase.table("emprestimos").update({"valor": novo_valor_total}).eq('id', emprestimo_id).execute()
            return response.data is not None
        except Exception as e:
            st.error(f"Erro ao atualizar o valor total do empréstimo: {e}")
            return False

    # Função para deletar um empréstimo e suas parcelas
    def delete_emprestimo(emprestimo_id):
        try:
            # Deletar parcelas relacionadas ao empréstimo
            supabase.table("parcelas").delete().eq("emprestimo_id", emprestimo_id).execute()
            # Deletar o empréstimo
            response = supabase.table("emprestimos").delete().eq('id', emprestimo_id).execute()
            return response.data is not None
        except Exception as e:
            st.error(f"Erro ao deletar empréstimo: {e}")
            return False

    # Página de Gerenciamento de Empréstimos
    page = st.radio("Selecione a página", ("Gerenciamento de Empréstimos", "Dashboard de Empréstimos"))

    if page == "Gerenciamento de Empréstimos":
        st.subheader("Gerenciamento de Empréstimos")

        emprestimos = load_emprestimos()

        # Criar duas colunas para os botões
        col1, col2 = st.columns(2)

        with col1:
            if st.button("Adicionar Novo Empréstimo"):
                st.session_state.show_create_form = True
                st.session_state.show_update_form = False
                st.session_state.show_delete_confirmation = False

        with col2:
            if st.button("Atualizar Empréstimo Existente"):
                if emprestimos:
                    st.session_state.show_update_form = True
                    st.session_state.show_create_form = False
                    st.session_state.show_delete_confirmation = False
                else:
                    st.warning("Nenhum empréstimo encontrado. Adicione um novo empréstimo primeiro.")

        with st.expander("Deletar Empréstimo", expanded=False):
            if emprestimos:
                emprestimo_selecionado = st.selectbox("Selecione um empréstimo para deletar", 
                                                        options=[e['nome'] for e in emprestimos], 
                                                        key=f"selectbox_deletar_emprestimos_{len(emprestimos)}")  # Chave única

                if st.button("Confirmar Deletar"):
                    existing_emprestimo = next((e for e in emprestimos if e['nome'] == emprestimo_selecionado), None)
                    if existing_emprestimo:
                        if delete_emprestimo(existing_emprestimo['id']):
                            st.success(f"Empréstimo '{existing_emprestimo['nome']}' deletado com sucesso!")
                        else:
                            st.error("Erro ao deletar o empréstimo.")
            else:
                st.warning("Nenhum empréstimo encontrado para deletar.")

        # Formulário para adicionar novo empréstimo
        if 'show_create_form' in st.session_state and st.session_state.show_create_form:
            st.write("Preencha os dados do novo empréstimo:")
            nome = st.text_input("Nome do Novo Empréstimo")
            valor = st.number_input("Valor do Empréstimo", min_value=0.0, step=0.01)
            data_inicial = st.date_input("Data Inicial")
            numero_parcelas = st.number_input("Número de Parcelas", min_value=1, step=1)

            # Adicione uma chave única ao checkbox
            if st.checkbox("Parcelas com valores personalizados?", key="checkbox_valores_personalizados"):
                valores_parcelas = []
                for i in range(int(numero_parcelas)):
                    valor_parcela = st.number_input(f"Valor da Parcela {i + 1}", min_value=0.0, step=0.01, key=f"input_parcela_{i}")
                    valores_parcelas.append(valor_parcela)
            else:
                valor_parcela = valor / numero_parcelas
                valores_parcelas = [valor_parcela] * int(numero_parcelas)

            if st.button("Salvar Novo Empréstimo"):
                emprestimo_existente = next((e for e in emprestimos if e['nome'].lower() == nome.lower()), None)
                if emprestimo_existente:
                    st.warning(f"Empréstimo com o nome '{nome}' já existe.")
                else:
                    # Verifique se os valores personalizados não ultrapassam o valor total do empréstimo
                    if st.session_state.get("checkbox_valores_personalizados", False):  # Verifica o valor do checkbox
                        total_valor_parcelas = sum(valores_parcelas)
                        if total_valor_parcelas > valor:
                            st.error("A soma das parcelas não pode ser maior que o valor total do empréstimo.")
                        else:
                            # Criação do objeto do empréstimo
                            emprestimo_data = {
                                "nome": nome,
                                "valor": valor,
                                "ano": data_inicial.year,
                                "mes": data_inicial.strftime("%b").lower(),
                                "situacao": "Em andamento",
                                "data_inicial": data_inicial.isoformat(),
                                "numero_parcelas": numero_parcelas,
                            }
                            response = save_emprestimo(emprestimo_data)
                            if response:
                                emprestimo_id = response[0]['id']
                                st.success("Novo empréstimo salvo com sucesso!")
                                parcelas_data = [
                                    {
                                        "emprestimo_id": emprestimo_id,
                                        "numero_parcela": i + 1,
                                        "mes_vencimento": (data_inicial + pd.DateOffset(months=i)).isoformat(),
                                        "valor_parcela": valores_parcelas[i],
                                        "status": "Não Pago",
                                    }
                                    for i in range(numero_parcelas)
                                ]
                                if save_parcelas(parcelas_data):
                                    st.success(f"{numero_parcelas} parcelas geradas com sucesso!")
                                else:
                                    st.error("Erro ao gerar as parcelas.")
                    else:
                        valor_parcela = valor / numero_parcelas
                        valores_parcelas = [valor_parcela] * int(numero_parcelas)

                        # Criação do objeto do empréstimo
                        emprestimo_data = {
                            "nome": nome,
                            "valor": valor,
                            "ano": data_inicial.year,
                            "mes": data_inicial.strftime("%b").lower(),
                            "situacao": "Em andamento",
                            "data_inicial": data_inicial.isoformat(),
                            "numero_parcelas": numero_parcelas,
                        }
                        response = save_emprestimo(emprestimo_data)
                        if response:
                            emprestimo_id = response[0]['id']
                            st.success("Novo empréstimo salvo com sucesso!")
                            parcelas_data = [
                                {
                                    "emprestimo_id": emprestimo_id,
                                    "numero_parcela": i + 1,
                                    "mes_vencimento": (data_inicial + pd.DateOffset(months=i)).isoformat(),
                                    "valor_parcela": valores_parcelas[i],
                                    "status": "Não Pago",
                                }
                                for i in range(numero_parcelas)
                            ]
                            if save_parcelas(parcelas_data):
                                st.success(f"{numero_parcelas} parcelas geradas com sucesso!")
                            else:
                                st.error("Erro ao gerar as parcelas.")
                        else:
                            st.error("Erro ao salvar o novo empréstimo.")

        # Formulário para atualizar o valor de uma parcela
        if 'show_update_form' in st.session_state and st.session_state.show_update_form:
            st.write("Atualizar Status da Parcela")
            
            # Escolher um empréstimo para atualizar
            emprestimo_selecionado_atualizacao = st.selectbox("Selecione um empréstimo para atualizar", 
                                                                options=[e['nome'] for e in emprestimos], 
                                                                key="selectbox_atualizar_emprestimo")

            existing_emprestimo = next((e for e in emprestimos if e['nome'] == emprestimo_selecionado_atualizacao), None)

            if existing_emprestimo:
                st.write(f"Gerenciar parcelas do empréstimo: {existing_emprestimo['nome']}")

                # Selecionar o status das parcelas a serem exibidas
                status_selecionado = st.selectbox("Escolha o status das parcelas:", ("Todos", "Pagas", "Não Pagas"), key="selectbox_status_parcelas")

                # Carregar as parcelas de acordo com o status selecionado
                parcelas = load_parcelas(existing_emprestimo['id'])
                if status_selecionado == "Pagas":
                    parcelas = [p for p in parcelas if p['status'] == "Pago"]
                elif status_selecionado == "Não Pagas":
                    parcelas = [p for p in parcelas if p['status'] == "Não Pago"]
                else:  # Para "Todos", mantém todas as parcelas
                    pass

                if parcelas:
                    # Selecionar uma parcela
                    parcela_selecionada = st.selectbox("Selecione uma parcela:", 
                                                        options=[f"Parcela {p['numero_parcela']}" for p in parcelas], 
                                                        key="selectbox_parcelas")
                    parcela = next((p for p in parcelas if f"Parcela {p['numero_parcela']}" == parcela_selecionada), None)

                    if parcela:
                        st.write(f"Você selecionou a parcela {parcela['numero_parcela']}.")

                        # Mostrar botão apenas se a parcela for "Não Pago"
                        if parcela['status'] == "Não Pago":
                            if st.button("Marcar como Pago"):
                                if marcar_parcela_como_pago(parcela['id']):
                                    st.success(f"Parcela {parcela['numero_parcela']} marcada como paga com sucesso!")
                                else:
                                    st.error("Erro ao marcar a parcela como paga.")
                        else:
                            st.warning("Essa parcela já está paga e não pode ser marcada novamente.")
                else:
                    st.warning("Nenhuma parcela encontrada com o status selecionado.")



    # Página do Dashboard de Empréstimos
    if page == "Dashboard de Empréstimos":
        st.subheader("Dashboard de Empréstimos")

        # Carregando os dados dos empréstimos
        emprestimos = load_emprestimos()

        # Verifica se existem empréstimos
        if emprestimos:
            total_emprestimos = len(emprestimos)
            total_valor = sum([e["valor"] for e in emprestimos])

            col1, col2 = st.columns(2)
            col1.metric("Total de Empréstimos", total_emprestimos)
            col2.metric("Valor Total de Empréstimos", f"R${total_valor:,.2f}")

            # Seletor para "Geral" ou um empréstimo específico
            emprestimo_selecionado = st.selectbox(
                "Selecione um empréstimo ou veja dados gerais",
                options=["Geral"] + [e["nome"] for e in emprestimos]
            )

            # Se for "Geral", exibe os dados gerais de todos os empréstimos
            if emprestimo_selecionado == "Geral":
                st.subheader("Dados Gerais de Todos os Empréstimos")

                # Cálculos dos totais gerais
                total_parcelas_pagas = 0
                total_parcelas_pendentes = 0
                total_valor_pago = 0
                total_valor_pendente = 0
                total_valor_vencido = 0

                for e in emprestimos:
                    parcelas = load_parcelas(e["id"])
                    for parcela in parcelas:
                        if parcela["status"] == "Pago" and parcela.get("valor_parcela") is not None:
                            total_parcelas_pagas += 1
                            total_valor_pago += parcela["valor_parcela"]
                        elif parcela["status"] == "Não Pago" and parcela.get("valor_parcela") is not None:
                            total_parcelas_pendentes += 1
                            total_valor_pendente += parcela["valor_parcela"]
                            if datetime.strptime(parcela["mes_vencimento"], '%Y-%m-%d') < datetime.now():
                                total_valor_vencido += parcela["valor_parcela"]

                # Exibindo os valores em texto
                st.subheader("Resumo Geral dos Empréstimos")
                st.write(f"**Total de Empréstimos:** {len(emprestimos)}")
                st.write(f"**Total de Parcelas Pagas:** {total_parcelas_pagas}")
                st.write(f"**Total de Parcelas Pendentes:** {total_parcelas_pendentes}")
                st.write(f"**Valor Total Pago:** R${total_valor_pago:,.2f}")
                st.write(f"**Valor Total Pendente:** R${total_valor_pendente:,.2f}")
                st.write(f"**Valor Total Vencido:** R${total_valor_vencido:,.2f}")

            else:
                # Detalhes de um empréstimo específico
                emprestimo = next(e for e in emprestimos if e["nome"] == emprestimo_selecionado)
                parcelas = load_parcelas(emprestimo["id"])

                st.subheader(f"Parcelas do Empréstimo: {emprestimo_selecionado}")
                df_parcelas = pd.DataFrame(parcelas)
                df_parcelas["mes_vencimento"] = pd.to_datetime(df_parcelas["mes_vencimento"])
                df_parcelas["status"] = df_parcelas["status"].apply(lambda x: "Pago" if x == "Pago" else "Não Pago")

                # Cálculos dos totais
                valor_total_emprestimo = emprestimo["valor"]
                parcelas_pagas = df_parcelas[df_parcelas["status"] == "Pago"].shape[0]
                parcelas_pendentes = df_parcelas[df_parcelas["status"] == "Não Pago"].shape[0]

                # Definindo parcelas atrasadas
                atrasadas = df_parcelas[(df_parcelas["status"] == "Não Pago") & (df_parcelas["mes_vencimento"] < pd.Timestamp.now())].shape[0]

                valor_pago = parcelas_pagas * (df_parcelas["valor_parcela"].iloc[0] if not df_parcelas["valor_parcela"].isnull().all() else 0)  # Verifica se não é None
                valor_pendente = parcelas_pendentes * (df_parcelas["valor_parcela"].iloc[0] if not df_parcelas["valor_parcela"].isnull().all() else 0)  # Verifica se não é None


                # Exibindo os valores em texto
                st.write(f"Valor Total do Empréstimo: R${valor_total_emprestimo:,.2f}")
                st.write(f"Total de Parcelas Pagas: {parcelas_pagas}")
                st.write(f"Total de Parcelas Pendentes: {parcelas_pendentes}")
                st.write(f"Valor Total Pago: R${valor_pago:,.2f}")
                st.write(f"Valor Total Pendente: R${valor_pendente:,.2f}")

                # Exibindo o gráfico
                fig = px.bar(
                    x=["Pagas", "Pendentes", "Atrasadas"],
                    y=[parcelas_pagas, parcelas_pendentes, atrasadas],
                    labels={"x": "Status", "y": "Quantidade de Parcelas"},
                    title="Distribuição de Parcelas Pagas, Pendentes e Atrasadas",
                    text=[parcelas_pagas, parcelas_pendentes, atrasadas]
                )
                st.plotly_chart(fig, use_container_width=True)
        else:
            st.warning("Nenhum empréstimo encontrado.")